from __future__ import annotations

import io
import json
import re
import shutil
import subprocess
import uuid
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any

import pdfplumber
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).parent
STORAGE_DIR = BASE_DIR / "storage"
UPLOADS_DIR = STORAGE_DIR / "uploads"
CONFIGS_DIR = STORAGE_DIR / "configs"

STORAGE_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
CONFIGS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Prospetti Ore & Costi API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class OcrMode(str, Enum):
    text_only = "TEXT_ONLY"
    text_then_ocr = "TEXT_THEN_OCR"
    ocr_force = "OCR_FORCE"


class ParsingConfig(BaseModel):
    name_regex: str = r"([A-Z][A-Z\s'`.-]{3,})"
    ore_ordinarie_regex: str = ""
    ore_straordinarie_regex: str = ""
    reperibilita_regex: str = ""
    netto_regex: str = ""
    pignoramento_regex: str = ""
    decimal_separator: str = ","
    thousands_separator: str = "."


class PageInfo(BaseModel):
    index: int
    text: str
    text_found: bool
    ocr_used: bool
    risk: str


class UploadResponse(BaseModel):
    upload_id: str
    filename: str
    pages: list[PageInfo]
    warnings: list[str]


class ResourceRow(BaseModel):
    id: str
    name: str
    role: str = ""
    ore_ordinarie: str = ""
    ore_straordinarie: str = ""
    reperibilita: str = ""
    netto: str = ""
    pignoramento: str = ""
    costo_orario: str = ""
    source: str = ""
    risk: str = ""


class ResourcesResponse(BaseModel):
    upload_id: str
    rows: list[ResourceRow]
    warnings: list[str] = []


class ConfigPayload(BaseModel):
    name: str
    data: dict[str, Any]


class ComputeRequest(BaseModel):
    upload_id: str
    config: dict[str, Any]
    resources: list[ResourceRow]


class ComputeResponse(BaseModel):
    upload_id: str
    allocations: list[dict[str, Any]]
    audit_log: list[dict[str, Any]]
    summary: dict[str, Any]
    warnings: list[str]


class ExportRequest(BaseModel):
    upload_id: str
    config: dict[str, Any]
    resources: list[ResourceRow]
    compute: dict[str, Any]


@app.get("/api/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/upload", response_model=UploadResponse)
async def upload_pdf(
    pdf: UploadFile = File(...),
    ocr_mode: OcrMode = Form(...),
    parsing_config: str | None = Form(None),
) -> UploadResponse:
    if not pdf.filename or not pdf.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo PDF consentiti")

    parsed_config = ParsingConfig()
    if parsing_config:
        try:
            parsed_config = ParsingConfig(**json.loads(parsing_config))
        except (json.JSONDecodeError, TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Invalid parsing_config JSON") from exc

    upload_id = str(uuid.uuid4())
    upload_dir = UPLOADS_DIR / upload_id
    upload_dir.mkdir(parents=True, exist_ok=True)

    source_path = upload_dir / "source.pdf"
    source_path.write_bytes(await pdf.read())

    pages, text_source, warnings = extract_pages(source_path, ocr_mode)
    resources, resource_warnings = extract_resources(pages, parsed_config, text_source)

    warnings.extend(resource_warnings)

    write_json(upload_dir / "pages.json", [page.model_dump() for page in pages])
    write_json(upload_dir / "resources.json", [row.model_dump() for row in resources])
    write_json(upload_dir / "parsing.json", parsed_config.model_dump())
    write_json(
        upload_dir / "meta.json",
        {
            "upload_id": upload_id,
            "filename": pdf.filename,
            "ocr_mode": ocr_mode.value,
            "created_at": datetime.utcnow().isoformat(),
        },
    )

    return UploadResponse(upload_id=upload_id, filename=pdf.filename, pages=pages, warnings=warnings)


@app.get("/api/pages")
async def get_pages(upload_id: str) -> JSONResponse:
    upload_dir = UPLOADS_DIR / upload_id
    pages_path = upload_dir / "pages.json"
    if not pages_path.exists():
        raise HTTPException(status_code=404, detail="Upload non trovato")
    pages = read_json(pages_path)
    return JSONResponse(content={"upload_id": upload_id, "pages": pages})


@app.get("/api/resources", response_model=ResourcesResponse)
async def get_resources(upload_id: str) -> ResourcesResponse:
    upload_dir = UPLOADS_DIR / upload_id
    resources_path = upload_dir / "resources.json"
    if not resources_path.exists():
        raise HTTPException(status_code=404, detail="Risorse non trovate")
    rows = [ResourceRow(**row) for row in read_json(resources_path)]
    return ResourcesResponse(upload_id=upload_id, rows=rows)


@app.post("/api/resources", response_model=ResourcesResponse)
async def save_resources(payload: ResourcesResponse) -> ResourcesResponse:
    upload_dir = UPLOADS_DIR / payload.upload_id
    if not upload_dir.exists():
        raise HTTPException(status_code=404, detail="Upload non trovato")
    resources_path = upload_dir / "resources.json"
    write_json(resources_path, [row.model_dump() for row in payload.rows])
    return payload


@app.post("/api/config")
async def save_config(payload: ConfigPayload) -> JSONResponse:
    config_path = CONFIGS_DIR / f"{slugify(payload.name)}.json"
    write_json(config_path, {"name": payload.name, "data": payload.data})
    return JSONResponse(content={"status": "ok", "path": str(config_path)})


@app.get("/api/config")
async def load_config(name: str) -> JSONResponse:
    config_path = CONFIGS_DIR / f"{slugify(name)}.json"
    if not config_path.exists():
        raise HTTPException(status_code=404, detail="Config non trovata")
    return JSONResponse(content=read_json(config_path))


@app.post("/api/compute", response_model=ComputeResponse)
async def compute_allocations(payload: ComputeRequest) -> ComputeResponse:
    warnings: list[str] = []
    config = payload.config
    consume_all = bool(config.get("consume_all_hours", False))

    networks = config.get("reti", [])
    roles = config.get("roles", [])
    people_rules = config.get("people_rules", {})

    if not networks or not roles:
        raise HTTPException(status_code=400, detail="Reti e ruoli sono obbligatori")

    resources = [row.model_dump() for row in payload.resources]
    allocations: list[dict[str, Any]] = []
    audit_log: list[dict[str, Any]] = []

    demand_by_role_network = build_demands(roles, networks, config, warnings)

    for person in resources:
        person_name = normalize_name(person.get("name", ""))
        if not person_name:
            continue
        hours_day = parse_number(person.get("ore_ordinarie")) + parse_number(
            person.get("ore_straordinarie")
        )
        hours_oncall = parse_number(person.get("reperibilita"))
        available_hours = {
            "day": hours_day,
            "oncall": hours_oncall,
        }
        person_roles = [person.get("role")] if person.get("role") else []
        if not person_roles:
            person_roles = [role.get("name") for role in roles]

        person_allocations = allocate_person_hours(
            person_name,
            person_roles,
            available_hours,
            demand_by_role_network,
            consume_all,
            roles,
            networks,
            audit_log,
            warnings,
        )
        allocations.extend(person_allocations)

    summary = summarize_allocations(demand_by_role_network, allocations, networks)

    return ComputeResponse(
        upload_id=payload.upload_id,
        allocations=allocations,
        audit_log=audit_log,
        summary=summary,
        warnings=warnings,
    )


@app.get("/api/export/excel")
async def export_excel(upload_id: str, config_name: str | None = None) -> StreamingResponse:
    upload_dir = UPLOADS_DIR / upload_id
    resources_path = upload_dir / "resources.json"
    if not resources_path.exists():
        raise HTTPException(status_code=404, detail="Risorse non trovate")

    resources = [ResourceRow(**row) for row in read_json(resources_path)]
    config_data: dict[str, Any] = {}
    if config_name:
        config_path = CONFIGS_DIR / f"{slugify(config_name)}.json"
        if config_path.exists():
            config_data = read_json(config_path).get("data", {})

    workbook = build_excel(resources, config_data)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "prospetti-ore-costi.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def extract_pages(source_path: Path, ocr_mode: OcrMode) -> tuple[list[PageInfo], str, list[str]]:
    warnings: list[str] = []
    pages_text = extract_text_per_page(source_path)
    text_found_ratio = sum(1 for text in pages_text if text.strip()) / max(len(pages_text), 1)

    ocr_used = False
    final_source = "text"

    if ocr_mode == OcrMode.text_only:
        if text_found_ratio == 0:
            warnings.append("PDF senza testo: OCR disattivato, nessuna estrazione valida.")
    else:
        needs_ocr = ocr_mode == OcrMode.ocr_force or text_found_ratio < 0.7
        if needs_ocr:
            ocr_path = source_path.parent / "ocr.pdf"
            if run_ocr(source_path, ocr_path):
                pages_text = extract_text_per_page(ocr_path)
                ocr_used = True
                final_source = "ocr"
            else:
                warnings.append("OCR non disponibile o fallito. Uso testo originale.")

    pages: list[PageInfo] = []
    for idx, text in enumerate(pages_text, start=1):
        risk = compute_risk(text, ocr_used)
        pages.append(
            PageInfo(
                index=idx,
                text=text,
                text_found=bool(text.strip()),
                ocr_used=ocr_used,
                risk=risk,
            )
        )
    return pages, final_source, warnings


def extract_resources(
    pages: list[PageInfo],
    config: ParsingConfig,
    text_source: str,
) -> tuple[list[ResourceRow], list[str]]:
    rows: list[ResourceRow] = []
    warnings: list[str] = []

    name_pattern = safe_compile(config.name_regex, "name_regex", warnings)
    ore_ord = safe_compile(config.ore_ordinarie_regex, "ore_ordinarie_regex", warnings)
    ore_str = safe_compile(config.ore_straordinarie_regex, "ore_straordinarie_regex", warnings)
    rep = safe_compile(config.reperibilita_regex, "reperibilita_regex", warnings)
    netto = safe_compile(config.netto_regex, "netto_regex", warnings)
    pignor = safe_compile(config.pignoramento_regex, "pignoramento_regex", warnings)

    for page in pages:
        names = []
        if name_pattern:
            names = list({match.group(1).strip() for match in name_pattern.finditer(page.text)})
        if not names:
            names = ["Nominativo da completare"]

        extracted = {
            "ore_ordinarie": find_value(ore_ord, page.text),
            "ore_straordinarie": find_value(ore_str, page.text),
            "reperibilita": find_value(rep, page.text),
            "netto": find_value(netto, page.text),
            "pignoramento": find_value(pignor, page.text),
        }
        risk = page.risk

        for name in names:
            rows.append(
                ResourceRow(
                    id=str(uuid.uuid4()),
                    name=normalize_name(name),
                    ore_ordinarie=extracted["ore_ordinarie"],
                    ore_straordinarie=extracted["ore_straordinarie"],
                    reperibilita=extracted["reperibilita"],
                    netto=extracted["netto"],
                    pignoramento=extracted["pignoramento"],
                    source=f"page:{page.index}|{text_source}",
                    risk=risk,
                )
            )

    return rows, warnings


def extract_text_per_page(path: Path) -> list[str]:
    with pdfplumber.open(path) as pdf:
        return [page.extract_text() or "" for page in pdf.pages]


def run_ocr(source: Path, output: Path) -> bool:
    if shutil.which("ocrmypdf") is None:
        return False
    try:
        subprocess.run(
            [
                "ocrmypdf",
                "--skip-text",
                "--optimize",
                "0",
                "--output-type",
                "pdf",
                str(source),
                str(output),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except subprocess.CalledProcessError:
        return False
    return output.exists()


def compute_risk(text: str, ocr_used: bool) -> str:
    if not text.strip():
        return "text_missing"
    if "�" in text or len(re.findall(r"[A-Za-z]", text)) < 5:
        return "low_quality"
    if "," in text and "." in text:
        return "ambiguous_number"
    if ocr_used:
        return "ocr_review"
    return "ok"


def safe_compile(regex: str, field: str, warnings: list[str]) -> re.Pattern | None:
    if not regex:
        return None
    try:
        return re.compile(regex)
    except re.error:
        warnings.append(f"Regex non valida per {field}: {regex}")
        return None


def find_value(pattern: re.Pattern | None, text: str) -> str:
    if not pattern:
        return ""
    match = pattern.search(text)
    if not match:
        return ""
    if match.groups():
        return match.group(1).strip()
    return match.group(0).strip()


def normalize_name(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", name).strip()
    return cleaned


def parse_number(value: str | None) -> float:
    if not value:
        return 0.0
    normalized = value.replace(".", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def build_demands(
    roles: list[dict[str, Any]],
    networks: list[str],
    config: dict[str, Any],
    warnings: list[str],
) -> dict[str, dict[str, float]]:
    demand_by_role_network: dict[str, dict[str, float]] = {}
    multipliers = config.get("period", {})
    days = multipliers.get("days", 1)
    weeks = multipliers.get("weeks", 1)
    nights = multipliers.get("nights", 1)

    for role in roles:
        role_name = role.get("name")
        if not role_name:
            continue
        demand = role.get("demand", {})
        demand_type = demand.get("type", "PER_MONTH")
        demand_value = float(demand.get("value", 0))
        if demand_type == "PER_DAY":
            total = demand_value * days
        elif demand_type == "PER_WEEK":
            total = demand_value * weeks
        elif demand_type == "PER_NIGHT":
            total = demand_value * nights
        elif demand_type == "FIXED_PER_RETE":
            total = demand_value
        else:
            total = demand_value

        demand_by_role_network[role_name] = {network: total for network in networks}

    return demand_by_role_network


def allocate_person_hours(
    person_name: str,
    roles_for_person: list[str],
    available_hours: dict[str, float],
    demand_by_role_network: dict[str, dict[str, float]],
    consume_all: bool,
    roles_config: list[dict[str, Any]],
    networks: list[str],
    audit_log: list[dict[str, Any]],
    warnings: list[str],
) -> list[dict[str, Any]]:
    allocations: list[dict[str, Any]] = []
    for role_name in roles_for_person:
        if role_name not in demand_by_role_network:
            continue
        role_cfg = next((role for role in roles_config if role.get("name") == role_name), {})
        chunk = float(role_cfg.get("allocation", {}).get("chunk", 1))
        allow_last = bool(role_cfg.get("allocation", {}).get("allow_last_fragment", True))
        last_step = float(role_cfg.get("allocation", {}).get("last_fragment_step", 0.5))
        remaining = available_hours.get("day", 0.0)

        while remaining > 0:
            target_network = pick_network(demand_by_role_network[role_name], consume_all)
            if not target_network:
                break
            assign = min(chunk, remaining)
            if assign < chunk and not allow_last:
                break
            if assign < chunk:
                assign = round_to_step(assign, last_step)
            remaining -= assign
            demand_by_role_network[role_name][target_network] = max(
                0.0, demand_by_role_network[role_name][target_network] - assign
            )
            allocations.append(
                {
                    "person": person_name,
                    "role": role_name,
                    "network": target_network,
                    "hours": assign,
                }
            )
            audit_log.append(
                {
                    "person": person_name,
                    "role": role_name,
                    "network": target_network,
                    "hours": assign,
                    "source": "auto",
                }
            )
            if not consume_all and demand_by_role_network[role_name][target_network] <= 0:
                break

        available_hours["day"] = remaining

    if consume_all and available_hours.get("day", 0.0) > 0:
        warnings.append(f"Ore residue non allocate per {person_name}")

    return allocations


def pick_network(demands: dict[str, float], consume_all: bool) -> str | None:
    if not demands:
        return None
    if consume_all:
        return min(demands.keys(), key=lambda key: demands[key])
    return max(demands.keys(), key=lambda key: demands[key])


def round_to_step(value: float, step: float) -> float:
    if step <= 0:
        return value
    return round(value / step) * step


def summarize_allocations(
    demand_by_role_network: dict[str, dict[str, float]],
    allocations: list[dict[str, Any]],
    networks: list[str],
) -> dict[str, Any]:
    totals = {network: 0.0 for network in networks}
    for entry in allocations:
        totals[entry["network"]] += entry["hours"]
    return {"allocated_hours": totals, "remaining_demand": demand_by_role_network}


def build_excel(resources: list[ResourceRow], config: dict[str, Any]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    networks = config.get("reti", []) or ["RETE_DEFAULT"]
    cigs = config.get("cig_groups", [])

    def sheet_name(prefix: str, name: str) -> str:
        return f"{prefix}{name}"[:31]

    headers = [
        "Nominativo",
        "Ruolo",
        "Ore ordinarie",
        "Ore straordinarie",
        "Reperibilita",
        "Netto",
        "Pignoramento",
    ]

    for network in networks:
        ws = workbook.create_sheet(title=sheet_name("RETE_", network))
        ws.append(headers)
        for row in resources:
            ws.append(
                [
                    row.name,
                    row.role,
                    row.ore_ordinarie,
                    row.ore_straordinarie,
                    row.reperibilita,
                    row.netto,
                    row.pignoramento,
                ]
            )
        ws.append([])
        ws.append(["FABBISOGNO"])
        ws.append(["ASSEGNATO"])
        ws.append(["DIFF"])
        ws.append(["CONTROLLO"])

    for cig in cigs:
        ws = workbook.create_sheet(title=sheet_name("CIG_", cig.get("name", "CIG")))
        ws.append(headers)
        for row in resources:
            ws.append(
                [
                    row.name,
                    row.role,
                    row.ore_ordinarie,
                    row.ore_straordinarie,
                    row.reperibilita,
                    row.netto,
                    row.pignoramento,
                ]
            )
        ws.append([])
        ws.append(["FABBISOGNO"])
        ws.append(["ASSEGNATO"])
        ws.append(["DIFF"])
        ws.append(["CONTROLLO"])

    ws_costi = workbook.create_sheet(title="Analisi_costi")
    ws_costi.append(["Nominativo", "Costo totale", "Note"])

    ws_controlli = workbook.create_sheet(title="Controlli")
    ws_controlli.append(["Check", "Esito", "Note"])

    ws_audit = workbook.create_sheet(title="Audit_Log")
    ws_audit.append(["Persona", "Ruolo", "Rete", "Ore", "Fonte"])

    return workbook


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-")
    return cleaned.lower() or "config"


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    try:
        import uvicorn
    except ImportError:
        raise SystemExit(
            "Uvicorn non installato. Installa con: pip install -r requirements.txt"
        )
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
