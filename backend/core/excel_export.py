from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook, load_workbook

from .models import AllocationRow


def build_consuntivo_excel(rows: List[AllocationRow], year: int, month: int) -> bytes:
    data = [
        {
            "Nominativo": row.name,
            "Rete": row.network,
            "Ruolo": row.role,
            "Ore": row.hours,
            "Costo_orario": row.cost_hour,
            "Importo": row.amount,
        }
        for row in rows
    ]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consuntivo")
    output.seek(0)
    return output.read()


def build_template_excel(
    rows: List[AllocationRow],
    year: int,
    month: int,
    template_path: Path | None,
) -> bytes:
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    networks = ["RETE1", "RETE2", "RETE3", "RETE4", "RETE5"]
    cig1_networks = set(networks[:4])

    def get_sheet(name: str):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(title=name)

    def write_sheet(sheet_name: str, filter_networks: set | None):
        ws = get_sheet(sheet_name)
        ws.delete_rows(1, ws.max_row)
        ws.append(["Nominativo", "Ruolo", "Ore", "Costo_orario", "Importo"])
        for row in rows:
            if filter_networks and row.network not in filter_networks:
                continue
            ws.append([row.name, row.role, row.hours, row.cost_hour, row.amount])
        ws.append([])
        ws.append(["FABBISOGNO (Ore)"])
        ws.append(["CONTROLLO COMMESSA"])

    write_sheet("CIG1", cig1_networks)
    for network in networks[:4]:
        write_sheet(network, {network})
    write_sheet("RETE5", {"RETE5"})

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def build_export_zip(
    rows: List[AllocationRow],
    year: int,
    month: int,
    template_path: Path | None,
) -> bytes:
    consuntivo = build_consuntivo_excel(rows, year, month)
    template = build_template_excel(rows, year, month, template_path)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        cons_name = f"PROSPETTO_CONSUNTIVO_{year}_{month:02d}.xlsx"
        template_name = f"CAS-PROSPETTO_ORE_FORMAT_TEMPLATE_{year}_{month:02d}_NOLOCK.xlsx"
        zf.writestr(cons_name, consuntivo)
        zf.writestr(template_name, template)

    zip_buffer.seek(0)
    return zip_buffer.read()
